"""
병원 경영진단 시스템 - 엑셀 입력 양식 생성기
실행: python create_template.py
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
import os

def make_border(thin=True):
    s = "thin" if thin else "medium"
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)

HEADER_FILL  = PatternFill("solid", fgColor="1565C0")
SUBHEAD_FILL = PatternFill("solid", fgColor="42A5F5")
SAMPLE_FILL  = PatternFill("solid", fgColor="E3F2FD")
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=11)
NORMAL_FONT  = Font(name="맑은 고딕", size=10)
BOLD_FONT    = Font(name="맑은 고딕", size=10, bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

MONTHS = [f"2024-{m:02d}" for m in range(1, 13)]
np.random.seed(42)


def style_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = CENTER_ALIGN
        cell.border = make_border()


def write_sheet_1_labor(wb):
    ws = wb.create_sheet("인건비")
    ws.column_dimensions["A"].width = 12
    for col in "BCDEF":
        ws.column_dimensions[col].width = 20

    # 헤더
    headers = ["월", "기본급합계(세전)", "4대보험(사용자부담)", "근로소득세",
               "인센티브", "퇴직충당금"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    # 샘플 데이터
    for m in MONTHS:
        row = [m,
               int(np.random.randint(28000000, 35000000)),
               int(np.random.randint(2500000,  3200000)),
               int(np.random.randint(800000,   1200000)),
               int(np.random.randint(0,        3000000)),
               int(np.random.randint(2200000,  2800000))]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col==1 else Alignment(horizontal="right", vertical="center")
            if col > 1:
                cell.number_format = "#,##0"

    ws.freeze_panes = "B2"
    return ws


def write_sheet_2_fixed(wb):
    ws = wb.create_sheet("고정비")
    headers = ["월","임차료","렌탈료","유지보수비",
               "전기료","수도료","통신비","보험료","기타고정비"]
    widths  = [12, 14, 12, 14, 10, 10, 10, 10, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    for m in MONTHS:
        row = [m, 4500000,
               int(np.random.randint(1200000, 1600000)),
               int(np.random.randint(300000,  800000)),
               int(np.random.randint(400000,  700000)),
               int(np.random.randint(80000,   150000)),
               int(np.random.randint(200000,  350000)),
               350000,
               int(np.random.randint(200000,  500000))]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col==1 else Alignment(horizontal="right", vertical="center")
            if col > 1:
                cell.number_format = "#,##0"
    ws.freeze_panes = "B2"
    return ws


def write_sheet_3_supply(wb):
    ws = wb.create_sheet("소모품약제")
    headers = ["월","의료소모품","약제비","장비구입비","위생소모품"]
    for i, w in enumerate([12,15,12,14,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    for m in MONTHS:
        row = [m,
               int(np.random.randint(1500000, 2500000)),
               int(np.random.randint(2000000, 3500000)),
               int(np.random.randint(0,       5000000)),
               int(np.random.randint(200000,  500000))]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col==1 else Alignment(horizontal="right", vertical="center")
            if col > 1:
                cell.number_format = "#,##0"
    ws.freeze_panes = "B2"
    return ws


def write_sheet_4_revenue(wb):
    ws = wb.create_sheet("매출")
    headers = ["월","급여매출","비급여매출","총매출","수납건수","1인당평균처방금액"]
    for i, w in enumerate([12,14,14,14,12,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    for idx, m in enumerate(MONTHS, 2):
        nhi    = int(np.random.randint(25000000, 40000000))
        non    = int(np.random.randint(8000000,  18000000))
        total  = nhi + non
        count  = int(np.random.randint(800, 1200))
        avg    = total // count
        row    = [m, nhi, non, total, count, avg]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col in (1,5) else Alignment(horizontal="right", vertical="center")
            if col in (2,3,4,6):
                cell.number_format = "#,##0"

    # 총계 행
    ws.append(["합계",
               f"=SUM(B2:B{len(MONTHS)+1})",
               f"=SUM(C2:C{len(MONTHS)+1})",
               f"=SUM(D2:D{len(MONTHS)+1})",
               f"=SUM(E2:E{len(MONTHS)+1})",
               f"=AVERAGE(F2:F{len(MONTHS)+1})"])
    total_row = ws.max_row
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill   = PatternFill("solid", fgColor="FFF3E0")
        cell.font   = BOLD_FONT
        cell.border = make_border(thin=False)
        cell.alignment = CENTER_ALIGN if col in (1,5) else Alignment(horizontal="right", vertical="center")
        if col in (2,3,4,6):
            cell.number_format = "#,##0"

    ws.freeze_panes = "B2"
    return ws


def write_sheet_5_outpatient(wb):
    ws = wb.create_sheet("원무")
    headers = ["월","신환수","총내원환자수","총내원횟수",
               "1인당평균내원횟수","주상병1","주상병2"]
    for i, w in enumerate([12,10,14,12,18,20,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    diseases1 = ["상세불명의 고혈압"]*4 + ["2형 당뇨병"]*4 + ["급성기관지염"]*4
    diseases2 = ["고지혈증"]*6 + ["요통"]*6

    for idx, m in enumerate(MONTHS):
        new_pt  = int(np.random.randint(60, 120))
        total_p = int(np.random.randint(700, 1100))
        total_v = int(np.random.randint(900, 1500))
        avg_v   = round(total_v / total_p, 2)
        row = [m, new_pt, total_p, total_v, avg_v, diseases1[idx], diseases2[idx]]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill      = SAMPLE_FILL
            cell.font      = NORMAL_FONT
            cell.border    = make_border()
            cell.alignment = CENTER_ALIGN if col in (1,5,6,7) else Alignment(horizontal="right", vertical="center")
    ws.freeze_panes = "B2"
    return ws


def write_sheet_6_staff(wb):
    ws = wb.create_sheet("직원현황")
    headers = ["직종","인원수","평균급여(세전)","월인건비합계","비고"]
    for i, w in enumerate([15,10,16,16,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    staff = [
        ("의사",      2, 8000000),
        ("간호사",    1, 3500000),
        ("간호조무사", 3, 2800000),
        ("원무",      2, 2600000),
        ("물리치료사", 1, 3000000),
        ("방사선사",  1, 3200000),
    ]
    for idx, (job, cnt, salary) in enumerate(staff, 2):
        total = cnt * salary
        ws.append([job, cnt, salary, total, ""])
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL if idx%2==0 else PatternFill("solid", fgColor="FFFFFF")
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col in (1,2,5) else Alignment(horizontal="right", vertical="center")
            if col in (3,4):
                cell.number_format = "#,##0"
    ws.freeze_panes = "A2"
    return ws


def write_sheet_7_marketing(wb):
    ws = wb.create_sheet("마케팅")
    headers = ["월","온라인광고비","오프라인광고비","SNS운영비","이벤트비용","합계"]
    for i, w in enumerate([12,14,15,12,12,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append(headers)
    style_header(ws, 1, len(headers))

    for idx, m in enumerate(MONTHS, 2):
        online  = int(np.random.randint(300000, 800000))
        offline = int(np.random.randint(0,      300000))
        sns     = int(np.random.randint(100000, 300000))
        event   = int(np.random.randint(0,      500000))
        total   = online + offline + sns + event
        row     = [m, online, offline, sns, event, total]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill   = SAMPLE_FILL
            cell.font   = NORMAL_FONT
            cell.border = make_border()
            cell.alignment = CENTER_ALIGN if col==1 else Alignment(horizontal="right", vertical="center")
            if col > 1:
                cell.number_format = "#,##0"
    ws.freeze_panes = "B2"
    return ws


def write_cover(wb):
    ws = wb.active
    ws.title = "📋 작성안내"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50

    # 타이틀
    ws.merge_cells("B1:C1")
    cell = ws["B1"]
    cell.value     = "🏥 병원 경영진단 시스템 - 데이터 입력 양식"
    cell.font      = Font(name="맑은 고딕", size=16, bold=True, color="1A237E")
    cell.alignment = CENTER_ALIGN
    cell.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:C2")
    ws["B2"].value     = "각 시트에 해당 데이터를 입력해 주세요. 샘플 데이터가 작성되어 있으니 참고하세요."
    ws["B2"].font      = Font(name="맑은 고딕", size=11, color="555555")
    ws["B2"].alignment = CENTER_ALIGN
    ws.row_dimensions[2].height = 25

    # 안내 테이블
    guide = [
        ("", "시트명", "입력 내용"),
        ("1", "인건비",    "월별 세전 인건비, 4대보험(사용자부담), 소득세, 인센티브, 퇴직충당금"),
        ("2", "고정비",    "월별 임차료, 렌탈료, 유지보수비, 전기·수도·통신·보험료"),
        ("3", "소모품약제", "월별 의료소모품, 약제비, 장비구입비, 위생소모품"),
        ("4", "매출",      "월별 급여매출, 비급여매출, 총매출, 수납건수"),
        ("5", "원무",      "월별 신환수, 총내원환자수, 총내원횟수, 주상병"),
        ("6", "직원현황",  "직종별 인원수, 평균급여"),
        ("7", "마케팅",    "월별 온라인/오프라인 광고비, SNS운영비, 이벤트비용"),
    ]
    for r_idx, row in enumerate(guide, 4):
        ws.row_dimensions[r_idx].height = 28
        for c_idx, val in enumerate(row, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = BOLD_FONT if r_idx==4 else NORMAL_FONT
            cell.border    = make_border()
            cell.alignment = CENTER_ALIGN
            if r_idx == 4:
                cell.fill = PatternFill("solid", fgColor="1565C0")
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            elif r_idx % 2 == 0:
                cell.fill = SAMPLE_FILL

    ws.merge_cells("B13:C13")
    ws["B13"].value     = "⚠️ 주의: 금액은 원(₩) 단위로 입력, 날짜는 YYYY-MM 형식으로 입력해 주세요."
    ws["B13"].font      = Font(name="맑은 고딕", size=10, color="C62828")
    ws["B13"].alignment = CENTER_ALIGN
    ws["B13"].fill      = PatternFill("solid", fgColor="FFEBEE")


def create_template(output_path="병원경영진단_입력양식.xlsx"):
    wb = Workbook()
    write_cover(wb)
    write_sheet_1_labor(wb)
    write_sheet_2_fixed(wb)
    write_sheet_3_supply(wb)
    write_sheet_4_revenue(wb)
    write_sheet_5_outpatient(wb)
    write_sheet_6_staff(wb)
    write_sheet_7_marketing(wb)
    wb.save(output_path)
    print(f"✅ 템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == "__main__":
    create_template()
